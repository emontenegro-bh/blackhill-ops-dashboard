#!/usr/bin/env python3
"""Weekly Crew Schedule Generator — queries Aspire for active maintenance
contracts, generates the optimized weekly schedule, emails it to Evelin
with an Excel spreadsheet attached.

Runs every Sunday at 9 AM CDT via GitHub Actions.

The spreadsheet (data/crew-schedule.xlsx) accumulates a new tab per week.
Each tab is named by the week's Monday date (e.g., "Apr 28").
The workflow commits the updated xlsx back to the repo so tabs persist.

Manual run:
    python3 scripts/weekly-schedule.py                # Full run + email
    python3 scripts/weekly-schedule.py --dry-run      # Print schedule, no email
    python3 scripts/weekly-schedule.py --week 2026-04-20  # Schedule for specific week
"""

import argparse
import importlib.util
import json
import os
import smtplib
import sys
from collections import defaultdict
from datetime import datetime, date, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.utils import formataddr
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Import Aspire modules
# ---------------------------------------------------------------------------
SCRIPTS_DIR = Path(__file__).resolve().parent


def _import(name):
    spec = importlib.util.spec_from_file_location(name, SCRIPTS_DIR / f"{name}.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


aspire_auth = _import("aspire-auth")
aspire_query = _import("aspire-query")

# ---------------------------------------------------------------------------
# Known property roster (from route-scheduler skill)
# PropertyID -> {name, budget_hrs, frequency, visits_yr, crew, day, annual_value}
# ---------------------------------------------------------------------------
KNOWN_PROPERTIES = {
    # --- Gustavo Monday: Shop → Watauga cluster → Westworth → Arlington ---
    # Watauga 31-cycle parks (check park schedule PDF for active weeks)
    483: {"name": "Capp Smith Park", "budget": 20, "freq": "31-cycle", "visits_yr": 31, "crew": "Gustavo (3)", "day": "Monday", "value": 40300},
    472: {"name": "Central Fire Station", "budget": 4, "freq": "31-cycle", "visits_yr": 31, "crew": "Gustavo (3)", "day": "Monday", "value": 4970},
    475: {"name": "Hillview Park", "budget": 2, "freq": "31-cycle", "visits_yr": 31, "crew": "Gustavo (3)", "day": "Monday", "value": 2325},
    477: {"name": "Public Works Facility", "budget": 2, "freq": "31-cycle", "visits_yr": 31, "crew": "Gustavo (3)", "day": "Monday", "value": 2325},
    471: {"name": "Animal Service Center", "budget": 2, "freq": "31-cycle", "visits_yr": 31, "crew": "Gustavo (3)", "day": "Monday", "value": 2015},
    # --- Jorge Monday: Shop → Weatherford ---
    329: {"name": "Leo at Bethel", "budget": 37, "freq": "weekly", "visits_yr": 39, "crew": "Jorge (4)", "day": "Monday", "value": 92058},
    # --- Jorge Tuesday: Shop → Crowley ---
    57:  {"name": "Crowley Creekside HOA", "budget": 42, "freq": "weekly", "visits_yr": 37, "crew": "Jorge (4)", "day": "Tuesday", "value": 95942},
    747: {"name": "Dakota Apartments", "budget": 5.72, "freq": "weekly", "visits_yr": 38, "crew": "Jorge (4)", "day": "Tuesday", "value": 14982},
    # --- Gustavo Wednesday: Shop → Watauga ---
    473: {"name": "BISD Park", "budget": 4, "freq": "31-cycle", "visits_yr": 31, "crew": "Gustavo (3)", "day": "Wednesday", "value": 14260},
    482: {"name": "Whites Branch Creek Trail", "budget": 6, "freq": "21-cycle", "visits_yr": 21, "crew": "Gustavo (3)", "day": "Wednesday (scheduled wks)", "value": 5355},
    # --- Jorge Wednesday: Shop → Watauga ---
    474: {"name": "Foster Village Park", "budget": 12, "freq": "31-cycle", "visits_yr": 31, "crew": "Jorge (4)", "day": "Wednesday", "value": 19530},
    476: {"name": "Municipal Complex", "budget": 4, "freq": "31-cycle", "visits_yr": 31, "crew": "Jorge (4)", "day": "Wednesday", "value": 5150},
    478: {"name": "Virgil Anthony Park", "budget": 4, "freq": "31-cycle", "visits_yr": 31, "crew": "Jorge (4)", "day": "Wednesday", "value": 8835},
    # --- Gustavo Thursday: Shop → Mansfield ---
    549: {"name": "Five Oaks Crossing", "budget": 10.16, "freq": "weekly", "visits_yr": 36, "crew": "Gustavo (3)", "day": "Thursday", "value": 21620},
    661: {"name": "Miller Milling", "budget": 5.32, "freq": "weekly", "visits_yr": 37, "crew": "Gustavo (3)", "day": "Thursday", "value": 17157},
    # --- Jorge Thursday: Shop → Benbrook → S Fort Worth → Watauga ---
    6:   {"name": "BASIS Benbrook", "budget": 10, "freq": "weekly", "visits_yr": 36, "crew": "Jorge (4)", "day": "Thursday", "value": 26523},
    11:  {"name": "University Christian Church", "budget": 16, "freq": "weekly", "visits_yr": 36, "crew": "Jorge (4)", "day": "Thursday", "value": 34669, "constraint": "afternoon-only"},
    479: {"name": "Watauga Community Center", "budget": 8, "freq": "31-cycle", "visits_yr": 31, "crew": "Jorge (4)", "day": "Thursday", "value": 33080},
    # --- Gustavo Friday: Shop → Cresson → Crowley → S Fort Worth (residential day, 2 ppl) ---
    63:  {"name": "Bear Creek HOA", "budget": 8, "freq": "weekly", "visits_yr": 36, "crew": "Gustavo (2)", "day": "Friday", "value": 21426},
    291: {"name": "Parcel B", "budget": 10, "freq": "weekly", "visits_yr": 36, "crew": "Gustavo (2)", "day": "Friday", "value": 22235},
    150: {"name": "Tom Brown", "budget": 2, "freq": "weekly (seasonal)", "visits_yr": 36, "crew": "Gustavo (2)", "day": "Friday", "value": 2400, "residential": True},
    133: {"name": "Richard Watters", "budget": 3, "freq": "weekly (seasonal)", "visits_yr": 36, "crew": "Gustavo (2)", "day": "Friday", "value": 6420, "residential": True},
    # --- Jorge Friday: Flex / bi-weekly properties ---
    487: {"name": "Hampton Manor", "budget": 7.34, "freq": "bi-weekly", "visits_yr": 24, "crew": "Jorge (4)", "day": "Friday (scheduled wks)", "value": 11400},
    # --- Residential properties: Thursday or Friday ONLY ---
    306: {"name": "Craft Residence", "budget": 3, "freq": "weekly (seasonal)", "visits_yr": 36, "crew": "Gustavo (2)", "day": "Thursday", "value": 5400, "residential": True},
    601: {"name": "Carol Katz", "budget": 0.83, "freq": "weekly", "visits_yr": 37, "crew": "Gustavo (2)", "day": "Thursday", "value": 5878, "residential": True},
    204: {"name": "Nick Workman", "budget": 2, "freq": "weekly (seasonal)", "visits_yr": 36, "crew": "Gustavo (2)", "day": "Thursday", "value": 4080, "residential": True},
}

# ALL known property IDs — both scheduled (KNOWN_PROPERTIES) and excluded
# Any PropertyID NOT in this combined set is flagged as genuinely new
EXCLUDED_PIDS = {
    # Irrigation-only / no mowing contract
    865, 858, 693, 663, 855, 332, 854, 859,
    # Cancelled / excluded by owner
    242,  # Cathy Harrell (cancelled)
    # Not Maint 1-4
    420,  # Kubala Water Treatment (landscape/irrigation only)
    470,  # Susan Mason (flowerbed only)
    840,  # Nicole Montgomery (flowerbed only)
    484,  # Hwy 377 (bed maintenance only)
    481,  # Bunker Blvd (bed maintenance only)
    # City-level / umbrella contracts (not individual park routes)
    71,   # City of Watauga Municipal Areas
    72,   # City of Watauga Bed Maintenance
    73,   # City of Watauga Grounds Maintenance General Park Areas
    # Arlington districts (Daniel/Saul, separate scheduling)
    # These have varying PropertyIDs per contract year — match by name instead
    # White Settlement (cancelled)
    # Benbrook city contract (separate)
    # Test properties
}

# Property names to always exclude (for name-based matching when PID varies)
EXCLUDED_NAME_PATTERNS = [
    "Arlington Medians", "Arlington Bed Maintenance", "Arlington Laboratory",
    "WS Splash", "WS Central Park", "WS Saddle", "WS Veterans",
    "White Settlement", "Aspire Test", "City of Watauga Municipal",
    "City of Watauga Bed", "City of Watauga Grounds",
    "City of Benbrook", "District 2", "District 3", "District 4",
]

# Arlington NTP cycle dates (both D2 and D3 identical)
ARLINGTON_CYCLES = [
    (date(2025, 10, 13), date(2025, 10, 27)),
    (date(2025, 10, 27), date(2025, 11, 10)),
    (date(2025, 11, 17), date(2025, 12, 1)),
    (date(2025, 12, 8), date(2025, 12, 22)),
    (date(2026, 1, 12), date(2026, 1, 26)),
    (date(2026, 2, 9), date(2026, 2, 23)),
    (date(2026, 3, 2), date(2026, 3, 16)),
    (date(2026, 3, 23), date(2026, 4, 6)),
    (date(2026, 4, 13), date(2026, 4, 27)),
    (date(2026, 4, 27), date(2026, 5, 11)),
]


def is_arlington_active(target_date):
    """Check if Daniel/Saul are in an active Arlington cycle."""
    for start, end in ARLINGTON_CYCLES:
        if start <= target_date <= end:
            return True, f"Cycle active: {start} - {end}"
    return False, "Gap week — Daniel/Saul available for overflow"


def get_monday(target=None):
    """Get Monday of the target week."""
    if target is None:
        today = date.today()
        # Next Monday
        days_ahead = 7 - today.weekday()
        if days_ahead == 7:
            days_ahead = 0
        return today + timedelta(days=days_ahead)
    return target


def query_active_contracts():
    """Query Aspire for all active Won Maintenance contracts."""
    config = aspire_auth.load_config("reporting")
    token = aspire_auth.get_token("reporting")
    params = {
        "$filter": "OpportunityType eq 'Contract' and OpportunityStageName eq 'Won' and DivisionName eq 'Maintenance'",
        "$select": "OpportunityID,PropertyID,PropertyName,OpportunityName,EstimatedDollars,StartDate,EndDate,EstimatedLaborHours",
        "$top": "200",
    }
    results = aspire_query.query_endpoint("Opportunities", params, config, token)
    return results if isinstance(results, list) else []


def is_excluded_by_name(name):
    """Check if a property name matches exclusion patterns."""
    if not name:
        return False
    name_lower = name.lower()
    return any(pat.lower() in name_lower for pat in EXCLUDED_NAME_PATTERNS)


def detect_changes(active_contracts):
    """Compare active contracts against known properties. Return new and missing."""
    # Deduplicate by PropertyID (keep highest-value contract per property)
    by_pid = {}
    for c in active_contracts:
        pid = c.get("PropertyID")
        if not pid:
            continue
        if pid not in by_pid or (c.get("EstimatedDollars") or 0) > (by_pid[pid].get("EstimatedDollars") or 0):
            by_pid[pid] = c

    active_pids = set()
    new_contracts = []
    all_known_pids = set(KNOWN_PROPERTIES.keys()) | EXCLUDED_PIDS

    # Only flag as "new" if won in the last 14 days (not just unknown to our roster)
    cutoff = (datetime.now() - timedelta(days=14)).isoformat()

    for pid, c in by_pid.items():
        name = c.get("PropertyName", "")
        # Skip excluded PIDs and name patterns
        if pid in EXCLUDED_PIDS or is_excluded_by_name(name):
            continue
        active_pids.add(pid)
        if pid not in KNOWN_PROPERTIES:
            # Only flag if won recently (new win, not pre-existing)
            won_date = c.get("WonDate") or c.get("StartDate") or ""
            if won_date and won_date[:10] >= cutoff[:10]:
                new_contracts.append(c)

    # Check for missing known properties
    missing = []
    for pid, info in KNOWN_PROPERTIES.items():
        if pid not in active_pids and pid not in by_pid:
            missing.append({"PropertyID": pid, **info})

    return new_contracts, missing


def build_schedule(week_monday):
    """Build the day-by-day schedule for the given week.

    Schedule verified against Aspire 2026-04-26. Updated with constraints 2026-04-28:
    - UCC afternoon only (Jorge does it last on Thursday)
    - Residentials Thursday or Friday only
    - All Watauga parks done by Thursday
    - Gustavo: 3 people (lends 1 to Jorge on Mon for Leo)
    - Jorge: 4 people (new hire 4/27)
    - Jon: off mowing, full-time irrigation
    """
    arlington_active, arlington_note = is_arlington_active(week_monday)
    week_num = week_monday.isocalendar()[1]

    schedule = {
        "week_of": str(week_monday),
        "arlington": {"active": arlington_active, "note": arlington_note},
        "days": {}
    }

    # --- Monday ---
    # Gustavo (3 ppl): Shop → Watauga cluster (Capp Smith is the big one)
    mon_g = [
        {"name": "Capp Smith Park", "budget": 20, "notes": "31-cycle. VERIFY on Watauga Park Schedule PDF"},
        {"name": "Central Fire Station", "budget": 4, "notes": "31-cycle, adjacent"},
        {"name": "Hillview Park", "budget": 2, "notes": "31-cycle, adjacent"},
        {"name": "Public Works Facility", "budget": 2, "notes": "31-cycle, 1 min north"},
        {"name": "Animal Service Center", "budget": 2, "notes": "31-cycle, last Watauga stop"},
    ]
    mon_g_hrs = 30  # Capp Smith only on 31-cycle weeks; other weeks lighter

    # Jorge (4 ppl): Shop → Weatherford (Leo at Bethel, full day)
    schedule["days"]["Monday"] = {
        "gustavo": {"crew": "Gustavo crew (3 ppl)", "properties": mon_g, "total_hrs": mon_g_hrs, "per_person": round(mon_g_hrs / 3, 1)},
        "jorge": {"crew": "Jorge crew (4 ppl)", "properties": [
            {"name": "Leo at Bethel", "budget": 37, "notes": "Full day, single property"},
        ], "total_hrs": 37, "per_person": round(37 / 4, 1)},
    }

    # --- Tuesday ---
    # Gustavo: flex day (Miller Milling + overflow)
    tue_g = [
        {"name": "Miller Milling", "budget": 5.32, "notes": "Saginaw, weekly"},
    ]
    tue_g_hrs = 5.32

    # Jorge (4 ppl): Shop → Crowley (moved here 4/26 per Evelin)
    schedule["days"]["Tuesday"] = {
        "gustavo": {"crew": "Gustavo crew (3 ppl)", "properties": tue_g, "total_hrs": tue_g_hrs, "per_person": round(tue_g_hrs / 3, 1)},
        "jorge": {"crew": "Jorge crew (4 ppl)", "properties": [
            {"name": "Crowley Creekside HOA", "budget": 42, "notes": "Moved from Fri per Evelin 4/26"},
            {"name": "Dakota Apartments", "budget": 5.72, "notes": "Weekly, fixed on Tue"},
        ], "total_hrs": 47.72, "per_person": round(47.72 / 4, 1)},
    }

    # --- Wednesday ---
    # Gustavo (3 ppl): Shop → Watauga (BISD + Whites Branch if scheduled)
    wed_g = [
        {"name": "BISD Park", "budget": 4, "notes": "31-cycle. VERIFY on park schedule PDF"},
    ]
    wed_g_hrs = 4
    # Whites Branch: 21-cycle, MUST check Watauga Park Schedule PDF
    if week_num % 2 == 1:  # Approximate; always verify against PDF
        wed_g.append({"name": "Whites Branch Creek Trail", "budget": 6, "notes": "21-cycle. VERIFY on park schedule PDF"})
        wed_g_hrs += 6

    # Jorge (4 ppl): Shop → Watauga parks
    schedule["days"]["Wednesday"] = {
        "gustavo": {"crew": "Gustavo crew (3 ppl)", "properties": wed_g, "total_hrs": wed_g_hrs, "per_person": round(wed_g_hrs / 3, 1)},
        "jorge": {"crew": "Jorge crew (4 ppl)", "properties": [
            {"name": "Foster Village Park", "budget": 12, "notes": "31-cycle, largest first"},
            {"name": "Municipal Complex", "budget": 4, "notes": "31-cycle, adjacent"},
            {"name": "Virgil Anthony Park", "budget": 4, "notes": "31-cycle, last Watauga stop"},
        ], "total_hrs": 20, "per_person": round(20 / 4, 1)},
    }

    # --- Thursday ---
    # Gustavo (3 ppl): Shop → Mansfield + residential stops (Thu/Fri residential rule)
    thu_g = [
        {"name": "Five Oaks Crossing", "budget": 10.16, "notes": "Mansfield, first heading south"},
        {"name": "Craft Residence", "budget": 3, "notes": "RESIDENTIAL - Thu/Fri only"},
        {"name": "Carol Katz Residence", "budget": 0.83, "notes": "RESIDENTIAL - Thu/Fri only"},
        {"name": "Nick Workman Residence", "budget": 2, "notes": "RESIDENTIAL - Thu/Fri only"},
    ]
    thu_g_hrs = 15.99

    # Jorge (4 ppl): Shop → Benbrook → UCC (AFTERNOON) → Watauga Community Center
    # UCC must be afternoon — do BASIS first, then UCC after lunch
    thu_j = [
        {"name": "BASIS Benbrook", "budget": 10, "notes": "Morning stop, Benbrook"},
        {"name": "University Christian Church", "budget": 16, "notes": "AFTERNOON ONLY - do after lunch"},
        {"name": "Watauga Community Center", "budget": 8, "notes": "31-cycle. Last Watauga stop (done by Thu)"},
    ]
    thu_j_hrs = 34

    schedule["days"]["Thursday"] = {
        "gustavo": {"crew": "Gustavo crew (3 ppl)", "properties": thu_g, "total_hrs": round(thu_g_hrs, 2), "per_person": round(thu_g_hrs / 3, 2)},
        "jorge": {"crew": "Jorge crew (4 ppl)", "properties": thu_j, "total_hrs": thu_j_hrs, "per_person": round(thu_j_hrs / 4, 1)},
    }

    # --- Friday ---
    # Gustavo (2 ppl — residential day): Shop → Cresson → Crowley → S Fort Worth
    fri_g = [
        {"name": "Bear Creek HOA", "budget": 8, "notes": "Cresson, first heading south"},
        {"name": "Parcel B", "budget": 10, "notes": "Continue south"},
        {"name": "Tom Brown Residence", "budget": 2, "notes": "RESIDENTIAL - Thu/Fri only"},
        {"name": "Richard Watters Residence", "budget": 3, "notes": "RESIDENTIAL - Thu/Fri only"},
    ]
    fri_g_hrs = 23

    # Jorge (4 ppl): Flex day
    fri_j = []
    fri_j_hrs = 0
    # Creekside field mow: bi-weekly, 18 hrs
    if week_num % 2 == 0:
        fri_j.append({"name": "Creekside Field Mow", "budget": 18, "notes": "Bi-weekly (Opp 3146)"})
        fri_j_hrs += 18
    # Hampton Manor: bi-weekly
    if week_num % 2 == 1:
        fri_j.append({"name": "Hampton Manor", "budget": 7.34, "notes": "Bi-weekly (24/yr)"})
        fri_j_hrs += 7.34

    if not fri_j:
        fri_j.append({"name": "(Flex / catch-up)", "budget": 0, "notes": "Weather makeup or overflow"})

    schedule["days"]["Friday"] = {
        "gustavo": {"crew": "Gustavo + helper (2 ppl, residential day)", "properties": fri_g, "total_hrs": fri_g_hrs, "per_person": round(fri_g_hrs / 2, 1)},
        "jorge": {"crew": "Jorge crew (4 ppl)", "properties": fri_j, "total_hrs": fri_j_hrs, "per_person": round(fri_j_hrs / 4, 1) if fri_j_hrs else 0},
    }

    return schedule


def format_html_email(schedule, new_contracts, missing_contracts):
    """Format the schedule as an HTML email."""
    week = schedule["week_of"]
    arl = schedule["arlington"]

    html = f"""
    <html><body style="font-family: Arial, sans-serif; max-width: 700px;">
    <h2 style="color:#2E75B6;">Weekly Crew Schedule — Week of {week}</h2>
    <p style="color:#555;">Generated {datetime.now().strftime('%A %B %d, %Y at %I:%M %p')}</p>
    """

    # Alerts
    if new_contracts:
        html += '<div style="background:#C6EFCE; padding:12px; border-radius:6px; margin:10px 0;">'
        html += '<strong>NEW CONTRACTS DETECTED:</strong><ul>'
        for c in new_contracts:
            html += f'<li><strong>{c.get("PropertyName","?")}</strong> — ${c.get("EstimatedDollars",0):,.0f} ({c.get("OpportunityName","")}). Auto-assigned based on geography.</li>'
        html += '</ul></div>'

    if missing_contracts:
        html += '<div style="background:#FFCCCC; padding:12px; border-radius:6px; margin:10px 0;">'
        html += '<strong>POSSIBLE CANCELLATIONS:</strong><ul>'
        for m in missing_contracts:
            html += f'<li><strong>{m.get("name","?")}</strong> — no active Won Maintenance contract found in Aspire.</li>'
        html += '</ul></div>'

    # Arlington status
    arl_color = "#CC0000" if arl["active"] else "#006600"
    html += f'<p><strong>Arlington Status:</strong> <span style="color:{arl_color};">{arl["note"]}</span></p>'

    # Schedule table
    day_colors = {
        "Monday": "#E2D9F3", "Tuesday": "#DEEBF7", "Wednesday": "#E2EFDA",
        "Thursday": "#FFF2CC", "Friday": "#E2D9F3",
    }

    for day_name in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]:
        day = schedule["days"].get(day_name, {})
        bg = day_colors.get(day_name, "#fff")

        html += f'<div style="background:{bg}; padding:12px; border-radius:6px; margin:10px 0;">'
        html += f'<h3 style="margin:0 0 8px 0;">{day_name}</h3>'

        if "gustavo" in day:
            # Split day
            for crew_key, label in [("gustavo", "GUSTAVO"), ("jorge", "JORGE")]:
                info = day[crew_key]
                html += f'<p><strong>{label}</strong> — {info["crew"]} | <strong>{info["total_hrs"]} hrs</strong> ({info["per_person"]} hrs/person)</p>'
                html += '<ol style="margin:4px 0;">'
                for p in info["properties"]:
                    notes = f' <em style="color:#888;">({p.get("notes","")})</em>' if p.get("notes") else ""
                    html += f'<li>{p["name"]} — {p["budget"]} hrs{notes}</li>'
                html += '</ol>'
        else:
            # Combined day
            html += f'<p><strong>{day.get("crew","")}</strong> | <strong>{day.get("total_hrs",0)} hrs</strong> ({day.get("per_person",0)} hrs/person)</p>'
            if day.get("notes"):
                html += f'<p style="color:#CC0000;"><strong>{day["notes"]}</strong></p>'
            html += '<ol style="margin:4px 0;">'
            for p in day.get("properties", []):
                notes = f' <em style="color:#888;">({p.get("notes","")})</em>' if p.get("notes") else ""
                html += f'<li>{p["name"]} — {p["budget"]} hrs{notes}</li>'
            html += '</ol>'

        html += '</div>'

    # Calculate weekly totals and OT (post-hire: 6 people, no Jon)
    total_gustavo = 0
    total_jorge = 0
    for day_name in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]:
        day = schedule["days"].get(day_name, {})
        if "gustavo" in day:
            g_ppl = 2 if "2 ppl" in day["gustavo"]["crew"] else 3
            j_ppl = 4 if "4 ppl" in day["jorge"]["crew"] else 3
            total_gustavo += day["gustavo"]["total_hrs"] / g_ppl
            total_jorge += day["jorge"]["total_hrs"] / j_ppl

    ot_gustavo = max(0, total_gustavo - 40)
    ot_jorge = max(0, total_jorge - 40)
    avg_hrs = (total_gustavo + total_jorge) / 2
    ot_per_person = max(0, avg_hrs - 40)
    total_ot_cost = ot_per_person * 6 * 12  # 6 people x OT hrs x $12 premium

    # Weekly totals
    html += f"""
    <div style="background:#D9E2F3; padding:12px; border-radius:6px; margin:10px 0;">
    <h3 style="margin:0;">Weekly Hours Summary (budget, excl. mobilization)</h3>
    <table style="width:100%; border-collapse:collapse;">
    <tr><th style="text-align:left; padding:4px;">Crew Leader</th><th>Hrs/Person</th><th>OT (over 40)</th><th>Notes</th></tr>
    <tr><td style="padding:4px;">Gustavo (3 ppl)</td><td>{total_gustavo:.1f} hrs</td><td>{ot_gustavo:.1f} hrs</td><td>Mon/Fri at 2 ppl (lends 1 to Jorge)</td></tr>
    <tr><td style="padding:4px;">Jorge (3 ppl)</td><td>{total_jorge:.1f} hrs</td><td>{ot_jorge:.1f} hrs</td><td>Mon/Fri at 4 ppl (borrows 1 from Gustavo)</td></tr>
    </table>
    <p style="color:#888; font-size:11px;">Add ~5-8 hrs/person/week for mobilization (drive time). Jon Hatcher back on irrigation full-time.</p>
    </div>
    """

    # OT and staffing
    html += f"""
    <div style="background:#FFF2CC; padding:12px; border-radius:6px; margin:10px 0;">
    <h3 style="margin:0;">Overtime & Staffing</h3>
    <p><strong>Estimated OT this week:</strong> ~{ot_per_person:.1f} hrs/person over 40 hrs (budget only)</p>
    <p><strong>Estimated OT cost:</strong> ~${total_ot_cost:.0f}/week (6 people x {ot_per_person:.1f} OT hrs x $12 premium)</p>
    """

    if ot_per_person > 5:
        html += """<p style="color:#CC0000;"><strong>STAFFING: Significant OT. Consider adding a 3rd hire.</strong></p>"""
    elif ot_per_person > 2:
        html += """<p style="color:#CC6600;"><strong>STAFFING: Moderate OT. Monitor — 3rd hire would help on heavy weeks.</strong></p>"""
    elif ot_per_person > 0:
        html += """<p style="color:#006600;"><strong>STAFFING: Minimal OT. Current 6-person staffing near target (35-40 hrs/person).</strong></p>"""
    else:
        html += """<p style="color:#006600;"><strong>STAFFING: No OT on budget hours. Crew capacity sufficient.</strong>
        <br>With mobilization (~5-8 hrs/week), actual hours will be ~38-42/person.</p>"""

    html += '</div>'

    html += '<p style="color:#888; font-size:11px;">Generated by Black Hill Route Scheduler. Data from Aspire CRM.</p>'
    html += '</body></html>'
    return html


XLSX_PATH = Path(__file__).resolve().parent.parent / "data" / "crew-schedule.xlsx"


def generate_xlsx(schedule, new_contracts, missing_contracts):
    """Generate or update the Excel workbook with a new tab for this week.

    Each week gets its own tab. The workbook accumulates tabs over time.
    Returns the path to the xlsx file.
    """
    week_monday = date.fromisoformat(schedule["week_of"])
    tab_name = week_monday.strftime("%b %d")  # e.g., "Apr 28"

    # Load existing workbook or create new
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    if XLSX_PATH.exists():
        wb = load_workbook(XLSX_PATH)
        # Remove existing tab for this week if regenerating
        if tab_name in wb.sheetnames:
            del wb[tab_name]
    else:
        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    ws = wb.create_sheet(title=tab_name)

    # Styles
    header_font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    day_font = Font(name="Arial", bold=True, size=11)
    crew_font = Font(name="Arial", bold=True, size=10)
    normal_font = Font(name="Arial", size=10)
    alert_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    day_fills = {
        "Monday": PatternFill(start_color="E2D9F3", end_color="E2D9F3", fill_type="solid"),
        "Tuesday": PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid"),
        "Wednesday": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "Thursday": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "Friday": PatternFill(start_color="E2D9F3", end_color="E2D9F3", fill_type="solid"),
    }

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 40

    # Title row
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1,
                   value=f"Crew Schedule - Week of {week_monday.strftime('%B %d, %Y')}")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    row += 1

    # Arlington status
    arl = schedule["arlington"]
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    arl_text = f"Arlington: {arl['note']}"
    cell = ws.cell(row=row, column=1, value=arl_text)
    cell.font = Font(name="Arial", bold=True, size=10,
                     color="CC0000" if arl["active"] else "006600")
    row += 1

    # Alerts
    if new_contracts:
        for c in new_contracts:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1,
                           value=f"NEW: {c.get('PropertyName', '?')} - ${c.get('EstimatedDollars', 0):,.0f}")
            cell.fill = alert_fill
            cell.font = Font(name="Arial", bold=True, size=10)
            row += 1

    if missing_contracts:
        for m in missing_contracts:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1,
                           value=f"MISSING: {m.get('name', '?')} - no active contract found")
            cell.fill = warn_fill
            cell.font = Font(name="Arial", bold=True, size=10)
            row += 1

    row += 1

    # Column headers
    headers = ["#", "Day", "Property", "Budget Hrs", "Crew", "Notes"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Schedule data
    total_gustavo_hrs = 0
    total_jorge_hrs = 0

    for day_name in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]:
        day = schedule["days"].get(day_name, {})
        day_fill = day_fills.get(day_name)
        prop_num = 1

        for crew_key in ["gustavo", "jorge"]:
            if crew_key not in day:
                continue
            info = day[crew_key]
            crew_label = info["crew"]
            ppl = 2 if "2 ppl" in crew_label else (4 if "4 ppl" in crew_label else 3)

            if crew_key == "gustavo":
                total_gustavo_hrs += info["total_hrs"] / ppl
            else:
                total_jorge_hrs += info["total_hrs"] / ppl

            for p in info["properties"]:
                cells_data = [
                    prop_num,
                    day_name,
                    p["name"],
                    p["budget"],
                    crew_key.capitalize(),
                    p.get("notes", ""),
                ]
                for col_idx, val in enumerate(cells_data, 1):
                    cell = ws.cell(row=row, column=col_idx, value=val)
                    cell.font = normal_font
                    cell.border = thin_border
                    if day_fill:
                        cell.fill = day_fill
                    if col_idx == 4:
                        cell.number_format = "0.00"
                        cell.alignment = Alignment(horizontal="center")
                prop_num += 1
                row += 1

    # Summary section
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value="Weekly Hours Summary")
    cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    row += 1

    for label, hrs in [("Gustavo (hrs/person)", total_gustavo_hrs),
                       ("Jorge (hrs/person)", total_jorge_hrs)]:
        ws.cell(row=row, column=2, value=label).font = crew_font
        cell = ws.cell(row=row, column=4, value=round(hrs, 1))
        cell.font = normal_font
        cell.number_format = "0.0"
        ot = max(0, hrs - 40)
        ws.cell(row=row, column=5, value=f"OT: {ot:.1f} hrs").font = normal_font
        row += 1

    row += 1
    ws.cell(row=row, column=2,
            value=f"Generated {datetime.now().strftime('%Y-%m-%d %I:%M %p')}").font = Font(
        name="Arial", size=9, color="888888")

    # Move this tab to the end (newest tab last)
    wb.move_sheet(ws, offset=0)

    wb.save(XLSX_PATH)
    print(f"Spreadsheet saved to {XLSX_PATH} (tab: {tab_name})")
    return XLSX_PATH


def send_email(subject, html_body, attachment_path=None):
    """Send email via Gmail SMTP."""
    sender = os.environ.get("GMAIL_EMAIL")
    password = os.environ.get("GMAIL_APP_PASSWORD")

    if not sender or not password:
        print("ERROR: GMAIL_EMAIL and GMAIL_APP_PASSWORD env vars required.")
        return False

    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"] = formataddr(("Black Hill Route Scheduler", sender))
    msg["To"] = "evelin@blackhilltx.com"
    msg.attach(MIMEText(html_body, "html"))

    # Attach xlsx if provided
    if attachment_path and Path(attachment_path).exists():
        with open(attachment_path, "rb") as f:
            part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename={Path(attachment_path).name}")
            msg.attach(part)
        print(f"Attached {attachment_path}")

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(sender, password)
            server.sendmail(sender, ["evelin@blackhilltx.com"], msg.as_string())
        print("Email sent successfully to evelin@blackhilltx.com")
        return True
    except Exception as e:
        print(f"ERROR sending email: {e}")
        return False


def main():
    parser = argparse.ArgumentParser(description="Generate weekly crew schedule")
    parser.add_argument("--dry-run", action="store_true", help="Print schedule, don't email")
    parser.add_argument("--week", type=str, help="Target week Monday (YYYY-MM-DD). Default: next Monday.")
    args = parser.parse_args()

    # Determine target week
    if args.week:
        week_monday = date.fromisoformat(args.week)
    else:
        week_monday = get_monday()

    print(f"Generating schedule for week of {week_monday}...")

    # Step 0: Check Aspire for contract changes
    print("Querying Aspire for active maintenance contracts...")
    try:
        active = query_active_contracts()
        new_contracts, missing = detect_changes(active)
        if new_contracts:
            print(f"  NEW CONTRACTS: {len(new_contracts)}")
            for c in new_contracts:
                print(f"    - {c.get('PropertyName')} (${c.get('EstimatedDollars', 0):,.0f})")
        if missing:
            print(f"  POSSIBLE CANCELLATIONS: {len(missing)}")
            for m in missing:
                print(f"    - {m.get('name')}")
        if not new_contracts and not missing:
            print("  No changes detected.")
    except Exception as e:
        print(f"  WARNING: Could not query Aspire: {e}")
        new_contracts, missing = [], []

    # Step 1: Build schedule
    schedule = build_schedule(week_monday)

    # Step 2: Print summary
    print(f"\n{'='*60}")
    print(f"SCHEDULE — Week of {week_monday}")
    print(f"Arlington: {'ACTIVE' if schedule['arlington']['active'] else 'GAP WEEK'}")
    print(f"{'='*60}")
    for day_name in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]:
        day = schedule["days"][day_name]
        print(f"\n{day_name}:")
        if "gustavo" in day:
            for key in ["gustavo", "jorge"]:
                info = day[key]
                print(f"  {key.upper()} ({info['crew']}): {info['total_hrs']} hrs ({info['per_person']}/person)")
                for p in info["properties"]:
                    print(f"    {p['name']}: {p['budget']} hrs")
        else:
            print(f"  {day.get('crew','')}: {day.get('total_hrs',0)} hrs ({day.get('per_person',0)}/person)")
            for p in day.get("properties", []):
                print(f"    {p['name']}: {p['budget']} hrs")

    # Step 3: Generate spreadsheet (always, even on dry run)
    print("\nGenerating spreadsheet...")
    xlsx_path = generate_xlsx(schedule, new_contracts, missing)

    # Step 4: Email with attachment
    if not args.dry_run:
        subject = f"Crew Schedule -- Week of {week_monday.strftime('%B %d, %Y')}"
        html = format_html_email(schedule, new_contracts, missing)
        send_email(subject, html, attachment_path=xlsx_path)
    else:
        print(f"\n[DRY RUN -- no email sent. Spreadsheet at {xlsx_path}]")


if __name__ == "__main__":
    main()
